"""Excel export: a multi-sheet workbook that stays *live*.

Totals, variances and percentages are written as **formulas**, not
Python-computed literals, so the workbook still adds up if the user edits a
figure in Excel. Every sheet carries a header row, currency number formats
matching the user's chosen currency, and a legend sheet explaining what is what.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

import config
from calculations.money import ZERO, D, money
from calculations.periods import Period
from constants import CURRENCY_FORMATS
from services import (
    account_service,
    budget_service,
    debt_service,
    goal_service,
    networth_service,
    reporting_service,
)
from services.common import category_name_map, settings_snapshot
from services.transaction_service import TxnFilter, list_transactions

FONT_NAME = "Arial"

HEADER_FILL = PatternFill("solid", fgColor="1F3B57")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=13, color="1F3B57")
BODY_FONT = Font(name=FONT_NAME, size=10)
TOTAL_FONT = Font(name=FONT_NAME, bold=True, size=10)
NOTE_FONT = Font(name=FONT_NAME, size=9, italic=True, color="595959")
INPUT_FONT = Font(name=FONT_NAME, size=10, color="0000FF")

THIN = Side(style="thin", color="D9D9D9")
BOX = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
TOP_RULE = Border(top=Side(style="thin", color="808080"))


def currency_format(currency: str) -> str:
    """Excel number format for the user's currency, negatives in parentheses."""
    fmt = CURRENCY_FORMATS.get((currency or "BRL").upper(), {})
    symbol = fmt.get("symbol", "")
    if (fmt.get("thousands"), fmt.get("decimal")) == (".", ","):
        # pt-BR style: Excel still uses the en-US pattern, the locale renders it.
        body = "#,##0.00"
    else:
        body = "#,##0.00"
    prefix = f'"{symbol} "' if symbol else ""
    return f'{prefix}{body};[Red]({prefix}{body});"-"'


PCT_FORMAT = "0.0%"
DATE_FORMAT = "yyyy-mm-dd"


@dataclass
class SheetSpec:
    title: str
    columns: list[tuple[str, str, int]]  # header, kind (text/money/pct/date/int), width


def _write_header(sheet: Worksheet, spec: SheetSpec, row: int = 1) -> int:
    for index, (header, _kind, width) in enumerate(spec.columns, start=1):
        cell = sheet.cell(row=row, column=index, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = sheet.cell(row=row + 1, column=1)
    sheet.auto_filter.ref = (
        f"A{row}:{get_column_letter(len(spec.columns))}{row}"
    )
    return row + 1


def _write_row(sheet: Worksheet, row: int, spec: SheetSpec,
               values: Sequence[Any], currency: str) -> None:
    money_fmt = currency_format(currency)
    for index, ((_header, kind, _width), value) in enumerate(
            zip(spec.columns, values), start=1):
        cell = sheet.cell(row=row, column=index, value=_coerce(value, kind))
        cell.font = BODY_FONT
        cell.border = BOX
        if kind == "money":
            cell.number_format = money_fmt
        elif kind == "pct":
            cell.number_format = PCT_FORMAT
        elif kind == "date":
            cell.number_format = DATE_FORMAT
        elif kind == "int":
            cell.number_format = "0"
        else:
            cell.alignment = Alignment(vertical="top", wrap_text=False)


def _coerce(value: Any, kind: str) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, str) and value.startswith("="):
        return value
    if kind == "money":
        return float(D(value))
    if kind == "pct":
        # Excel stores percentages as fractions.
        return float(D(value)) / 100.0
    if kind == "int":
        try:
            return int(value)
        except (TypeError, ValueError):
            return None
    if kind == "date":
        if isinstance(value, datetime):
            return value.date()
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, bool):
        return "yes" if value else "no"
    return value


def _total_row(sheet: Worksheet, row: int, spec: SheetSpec, first_data_row: int,
               currency: str, label: str = "TOTAL") -> None:
    money_fmt = currency_format(currency)
    last = row - 1
    sheet.cell(row=row, column=1, value=label).font = TOTAL_FONT
    for index, (_header, kind, _width) in enumerate(spec.columns, start=1):
        cell = sheet.cell(row=row, column=index)
        cell.border = TOP_RULE
        if index == 1:
            continue
        if kind == "money" and last >= first_data_row:
            letter = get_column_letter(index)
            cell.value = f"=SUM({letter}{first_data_row}:{letter}{last})"
            cell.number_format = money_fmt
            cell.font = TOTAL_FONT


# ==========================================================================
# Sheets
# ==========================================================================
TXN_SPEC = SheetSpec("Transactions", [
    ("Date", "date", 12), ("Paid on", "date", 12), ("Description", "text", 40),
    ("Type", "text", 11), ("Status", "text", 11), ("Category", "text", 28),
    ("Account", "text", 22), ("To account", "text", 20), ("Amount", "money", 15),
    ("Payment method", "text", 16), ("Tags", "text", 16), ("Planned", "text", 9),
    ("Notes", "text", 30),
])

BUDGET_SPEC = SheetSpec("Budget", [
    ("Period", "text", 11), ("Type", "text", 12), ("Line", "text", 34),
    ("Planned", "money", 15), ("Actual", "money", 15), ("Variance", "money", 15),
    ("Variance %", "pct", 12), ("Consumed %", "pct", 12), ("Status", "text", 18),
])

HISTORY_SPEC = SheetSpec("History", [
    ("Period", "text", 11), ("Income", "money", 15), ("Expenses", "money", 15),
    ("Savings", "money", 15), ("Investments", "money", 15),
    ("Debt payments", "money", 15), ("Net", "money", 15),
    ("Closing cash", "money", 16), ("Savings rate", "pct", 12),
])

ACCOUNT_SPEC = SheetSpec("Accounts", [
    ("Account", "text", 26), ("Type", "text", 22), ("Currency", "text", 9),
    ("Balance", "money", 16), ("Owed", "money", 16), ("Credit limit", "money", 14),
    ("Utilisation", "pct", 12), ("In net worth", "text", 12),
])

GOAL_SPEC = SheetSpec("Goals", [
    ("Goal", "text", 28), ("Type", "text", 20), ("Target", "money", 15),
    ("Saved", "money", 15), ("Remaining", "money", 15), ("Progress", "pct", 11),
    ("Planned monthly", "money", 16), ("Required monthly", "money", 16),
    ("Target date", "date", 13), ("Projected finish", "date", 15),
    ("On track", "text", 10),
])

DEBT_SPEC = SheetSpec("Debts", [
    ("Debt", "text", 28), ("Type", "text", 18), ("Balance", "money", 15),
    ("Annual rate %", "text", 13), ("Minimum", "money", 14),
    ("Planned payment", "money", 16), ("Extra", "money", 12),
    ("Monthly interest", "money", 16), ("Months to payoff", "int", 15),
    ("Total interest", "money", 16), ("Payoff date", "date", 13),
])

NETWORTH_SPEC = SheetSpec("Net worth", [
    ("As of", "date", 12), ("Assets", "money", 16),
    ("Liabilities", "money", 16), ("Net worth", "money", 16),
])

FORECAST_SPEC = SheetSpec("Forecast", [
    ("Period", "text", 11), ("Basis", "text", 22), ("Income", "money", 15),
    ("Expenses", "money", 15), ("Savings", "money", 15),
    ("Investments", "money", 15), ("Debt payments", "money", 15),
    ("Net flow", "money", 15), ("Closing cash", "money", 16),
    ("Free of earmarks", "money", 17),
])


def build_workbook(
    session: Session,
    *,
    months: int = 12,
    include_forecast: bool = True,
    today: Optional[date] = None,
) -> Workbook:
    """Assemble the full export."""
    settings = settings_snapshot(session)
    today = today or date.today()
    currency = settings.base_currency
    periods = reporting_service.trailing_periods(session, months, today)

    workbook = Workbook()
    _legend_sheet(workbook, settings, today, months)
    _transactions_sheet(workbook, session, periods, currency)
    _budget_sheet(workbook, session, periods, currency)
    _history_sheet(workbook, session, periods, currency, today)
    _accounts_sheet(workbook, session, currency)
    _goals_sheet(workbook, session, currency, today)
    _debts_sheet(workbook, session, currency, today)
    _networth_sheet(workbook, session, currency, months, today)
    if include_forecast:
        _forecast_sheet(workbook, session, currency, settings, today)
    return workbook


def _legend_sheet(workbook: Workbook, settings, today: date, months: int) -> None:
    sheet = workbook.active
    sheet.title = "About"
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 78

    sheet["A1"] = f"{config.APP_NAME} — data export"
    sheet["A1"].font = TITLE_FONT

    lines = [
        ("Exported", today.isoformat()),
        ("Currency", settings.base_currency),
        ("Period grain", f"Monthly, starting on day {settings.first_day_of_month}"),
        ("Range", f"Last {months} period(s) up to {today.isoformat()}"),
        ("Income availability", settings.income_availability_rule.replace("_", " ")),
        ("", ""),
        ("Transactions", "Every recorded and planned movement. Amounts are always "
                         "positive; direction comes from the Type column."),
        ("Budget", "Planned versus actual by budget line. Variance = Actual − Planned. "
                   "For expenses a negative variance is good; for income a positive one is."),
        ("History", "One row per period with the headline figures and closing cash."),
        ("Accounts", "Current balances. Liabilities are shown in the Owed column as a "
                     "positive amount."),
        ("Goals", "Progress, and the monthly contribution required to hit the target date."),
        ("Debts", "Payoff projection at the current payment, interest compounded monthly."),
        ("Net worth", "Assets minus liabilities at the end of each period."),
        ("Forecast", "Forward projection. 'Basis' says where each period's numbers came "
                     "from: your budget, your recurring rules, or an average of history."),
        ("", ""),
        ("Formulas", "Total rows are live SUM formulas, so the workbook still adds up "
                     "if you edit a figure."),
        ("Blue text", "Values you may want to change; everything else is exported data."),
    ]
    row = 3
    for label, text in lines:
        if label:
            cell = sheet.cell(row=row, column=1, value=label)
            cell.font = TOTAL_FONT if not text else BODY_FONT
        body = sheet.cell(row=row, column=2, value=text)
        body.font = BODY_FONT
        body.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    note = sheet.cell(row=row + 1, column=2,
                      value="All figures come from your local database; nothing was "
                            "sent anywhere to produce this file.")
    note.font = NOTE_FONT


def _transactions_sheet(workbook: Workbook, session: Session,
                        periods: Sequence[Period], currency: str) -> None:
    sheet = workbook.create_sheet(TXN_SPEC.title)
    row = _write_header(sheet, TXN_SPEC)
    first = row
    if not periods:
        return
    categories = category_name_map(session)
    accounts = {a.id: a.name for a in account_service.list_accounts(
        session, include_archived=True)}
    txns = list_transactions(session, TxnFilter(
        start=periods[0].start, end=periods[-1].end, order_desc=False,
    ))
    for txn in txns:
        _write_row(sheet, row, TXN_SPEC, [
            txn.txn_date, txn.actual_date, txn.description, txn.kind, txn.status,
            categories.get(txn.category_id, "") if txn.category_id else "",
            accounts.get(txn.account_id, "") if txn.account_id else "",
            accounts.get(txn.to_account_id, "") if txn.to_account_id else "",
            txn.amount, txn.payment_method or "", txn.tags or "",
            "yes" if txn.is_planned else "no", txn.notes or "",
        ], currency)
        row += 1
    if row > first:
        _total_row(sheet, row, TXN_SPEC, first, currency)


def _budget_sheet(workbook: Workbook, session: Session,
                  periods: Sequence[Period], currency: str) -> None:
    sheet = workbook.create_sheet(BUDGET_SPEC.title)
    row = _write_header(sheet, BUDGET_SPEC)
    first = row
    for period in periods:
        tracking = budget_service.track_period(session, period)
        for entry in tracking.rows:
            variance_formula = f"=E{row}-D{row}"
            variance_pct = f'=IF(D{row}=0,"",(E{row}-D{row})/D{row})'
            consumed = f'=IF(D{row}=0,"",E{row}/D{row})'
            _write_row(sheet, row, BUDGET_SPEC, [
                period.key, entry.kind, entry.label,
                entry.planned, entry.actual,
                variance_formula, variance_pct, consumed,
                f"{entry.status_icon} {entry.status_label}",
            ], currency)
            row += 1
    if row > first:
        _total_row(sheet, row, BUDGET_SPEC, first, currency)


def _history_sheet(workbook: Workbook, session: Session, periods: Sequence[Period],
                   currency: str, today: date) -> None:
    sheet = workbook.create_sheet(HISTORY_SPEC.title)
    row = _write_header(sheet, HISTORY_SPEC)
    first = row
    for entry in reporting_service.period_history(session, periods, today=today):
        net_formula = f"=B{row}-C{row}-D{row}-E{row}-F{row}"
        rate_formula = f'=IF(B{row}=0,"",(D{row}+E{row})/B{row})'
        _write_row(sheet, row, HISTORY_SPEC, [
            entry["label"], entry["income"], entry["expenses"], entry["savings"],
            entry["investments"], entry["debt_payments"], net_formula,
            entry["closing_cash"], rate_formula,
        ], currency)
        row += 1
    if row > first:
        _total_row(sheet, row, HISTORY_SPEC, first, currency)


def _accounts_sheet(workbook: Workbook, session: Session, currency: str) -> None:
    sheet = workbook.create_sheet(ACCOUNT_SPEC.title)
    row = _write_header(sheet, ACCOUNT_SPEC)
    first = row
    for view in account_service.balance_views(session, include_archived=True):
        _write_row(sheet, row, ACCOUNT_SPEC, [
            view.name, view.type_label, view.account.currency,
            view.balance if not view.is_liability else ZERO,
            view.display_balance if view.is_liability else ZERO,
            view.account.credit_limit or ZERO,
            view.utilisation_pct if view.utilisation_pct is not None else None,
            "yes" if view.account.include_in_net_worth else "no",
        ], currency)
        row += 1
    if row > first:
        _total_row(sheet, row, ACCOUNT_SPEC, first, currency)


def _goals_sheet(workbook: Workbook, session: Session, currency: str, today: date) -> None:
    sheet = workbook.create_sheet(GOAL_SPEC.title)
    row = _write_header(sheet, GOAL_SPEC)
    first = row
    from constants import GOAL_TYPE_LABELS, GoalStatus

    for goal in goal_service.list_goals(session):
        progress = goal_service.progress_for(session, goal, today=today)
        remaining_formula = f"=MAX(0,C{row}-D{row})"
        progress_formula = f'=IF(C{row}=0,"",MIN(1,D{row}/C{row}))'
        _write_row(sheet, row, GOAL_SPEC, [
            goal.name, GOAL_TYPE_LABELS.get(goal.goal_type, goal.goal_type),
            goal.target_amount, progress.current_amount, remaining_formula,
            progress_formula, goal.planned_monthly, progress.required_monthly,
            goal.target_date, progress.projected_completion,
            "yes" if progress.on_track else ("no" if progress.on_track is False else ""),
        ], currency)
        row += 1
    if row > first:
        _total_row(sheet, row, GOAL_SPEC, first, currency)


def _debts_sheet(workbook: Workbook, session: Session, currency: str, today: date) -> None:
    sheet = workbook.create_sheet(DEBT_SPEC.title)
    row = _write_header(sheet, DEBT_SPEC)
    first = row
    from constants import DEBT_TYPE_LABELS

    for view in debt_service.views(session, today=today):
        projection = view.projection
        _write_row(sheet, row, DEBT_SPEC, [
            view.debt.name,
            DEBT_TYPE_LABELS.get(view.debt.debt_type, view.debt.debt_type),
            view.balance, f"{D(view.debt.interest_rate):.2f}%",
            view.debt.minimum_payment, view.debt.planned_payment,
            view.debt.extra_payment, view.monthly_interest,
            projection.months if not projection.never_pays_off else None,
            projection.total_interest if not projection.never_pays_off else None,
            projection.payoff_date if not projection.never_pays_off else None,
        ], currency)
        if projection.never_pays_off:
            sheet.cell(row=row, column=9, value="never at this payment").font = NOTE_FONT
        row += 1
    if row > first:
        _total_row(sheet, row, DEBT_SPEC, first, currency)


def _networth_sheet(workbook: Workbook, session: Session, currency: str,
                    months: int, today: date) -> None:
    sheet = workbook.create_sheet(NETWORTH_SPEC.title)
    row = _write_header(sheet, NETWORTH_SPEC)
    for point in networth_service.trailing_history(session, months, today):
        net_formula = f"=B{row}-C{row}"
        _write_row(sheet, row, NETWORTH_SPEC, [
            point.as_of, point.total_assets, point.total_liabilities, net_formula,
        ], currency)
        row += 1


def _forecast_sheet(workbook: Workbook, session: Session, currency: str,
                    settings, today: date) -> None:
    from services import forecast_service

    sheet = workbook.create_sheet(FORECAST_SPEC.title)
    row = _write_header(sheet, FORECAST_SPEC)
    first = row
    bundle = forecast_service.build(
        session, months=settings.forecast_months, history_months=0, today=today
    )
    for entry in bundle.rows:
        assumption = entry.assumption
        savings = money(assumption.savings_reserved + assumption.savings_outflow)
        net_formula = f"=C{row}-D{row}-F{row}-G{row}"
        _write_row(sheet, row, FORECAST_SPEC, [
            entry.period.key, entry.source_label, assumption.income,
            assumption.expenses, savings, assumption.investments,
            assumption.debt_payments, net_formula, entry.closing_cash,
            entry.free_cash,
        ], currency)
        row += 1
    if row > first:
        note = sheet.cell(row=row + 1, column=2,
                          value="Net flow excludes savings that stay inside a cash account "
                                "— that money is earmarked, not spent.")
        note.font = NOTE_FONT


def workbook_bytes(session: Session, **kwargs) -> bytes:
    """The workbook as bytes, ready for a Streamlit download button."""
    workbook = build_workbook(session, **kwargs)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def save_workbook(session: Session, path, **kwargs):
    from pathlib import Path

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(session, **kwargs).save(target)
    return target
